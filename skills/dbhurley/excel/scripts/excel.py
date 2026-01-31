#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel 命令行工具 - 全面的 Excel 文件操作。

读取、写入、编辑和格式化 Excel 文件 (.xlsx, .xls)。

使用方法:
    excel.py info <file>
    excel.py read <file> [--sheet NAME] [--range A1:B10] [--format json|csv|markdown]
    excel.py cell <file> <cell> [--sheet NAME]
    excel.py create <file> [--sheets NAME,NAME2]
    excel.py write <file> --data JSON [--sheet NAME] [--start A1]
    excel.py from-csv <csv_file> <excel_file> [--sheet NAME]
    excel.py from-json <json_file> <excel_file> [--sheet NAME]
    excel.py edit <file> <cell> <value> [--sheet NAME] [--formula]
    excel.py add-sheet <file> <name> [--position N]
    excel.py rename-sheet <file> <old_name> <new_name>
    excel.py delete-sheet <file> <name>
    excel.py copy-sheet <file> <source> <new_name>
    excel.py insert-rows <file> <row> [--count N] [--sheet NAME]
    excel.py insert-cols <file> <col> [--count N] [--sheet NAME]
    excel.py delete-rows <file> <row> [--count N] [--sheet NAME]
    excel.py delete-cols <file> <col> [--count N] [--sheet NAME]
    excel.py merge <file> <range> [--sheet NAME]
    excel.py unmerge <file> <range> [--sheet NAME]
    excel.py format <file> <range> [--sheet NAME] [options...]
    excel.py resize <file> [--row N:HEIGHT] [--col A:WIDTH] [--sheet NAME]
    excel.py freeze <file> <cell> [--sheet NAME]
    excel.py find <file> <text> [--sheet NAME]
    excel.py replace <file> <old> <new> [--sheet NAME]
    excel.py to-csv <file> <output> [--sheet NAME]
    excel.py to-json <file> <output> [--sheet NAME]
    excel.py to-markdown <file> [--sheet NAME]
"""

import argparse
import csv
import json
import os
import re
import sys
from pathlib import Path
from typing import Any, Optional, Union

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.utils import get_column_letter, column_index_from_string
    from openpyxl.utils.cell import coordinate_from_string
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.styles.colors import Color
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    import xlrd
    XLRD_AVAILABLE = True
except ImportError:
    XLRD_AVAILABLE = False


def require_openpyxl():
    if not OPENPYXL_AVAILABLE:
        print("Error: openpyxl required. Install with: pip install openpyxl", file=sys.stderr)
        sys.exit(1)


def ok(data: Any):
    """输出成功 JSON。"""
    print(json.dumps({"success": True, **data}, indent=2, default=str))


def fail(message: str, details: Any = None):
    """输出错误 JSON。"""
    result = {"success": False, "error": message}
    if details:
        result["details"] = details
    print(json.dumps(result, indent=2, default=str))
    sys.exit(1)


def parse_range(range_str: str) -> tuple:
    """解析 A1:B10 为 ((1,1), (10,2))。"""
    if ':' in range_str:
        start, end = range_str.upper().split(':')
    else:
        start = end = range_str.upper()
    
    start_col, start_row = coordinate_from_string(start)
    end_col, end_row = coordinate_from_string(end)
    
    return (
        (start_row, column_index_from_string(start_col)),
        (end_row, column_index_from_string(end_col))
    )


def cell_to_coords(cell: str) -> tuple:
    """将 A1 转换为 (行, 列) 元组。"""
    col_str, row = coordinate_from_string(cell.upper())
    return (row, column_index_from_string(col_str))


def coords_to_cell(row: int, col: int) -> str:
    """将 (行, 列) 转换为 A1 表示法。"""
    return f"{get_column_letter(col)}{row}"


def get_sheet(wb, sheet_name: Optional[str] = None):
    """通过名称获取工作表或活动工作表。"""
    if sheet_name:
        if sheet_name not in wb.sheetnames:
            fail(f"Sheet '{sheet_name}' not found", {"available": wb.sheetnames})
        return wb[sheet_name]
    return wb.active


def read_sheet_data(ws, range_str: Optional[str] = None) -> list:
    """将工作表数据读取为列表的列表。"""
    if range_str:
        (start_row, start_col), (end_row, end_col) = parse_range(range_str)
        data = []
        for row in range(start_row, end_row + 1):
            row_data = []
            for col in range(start_col, end_col + 1):
                cell = ws.cell(row=row, column=col)
                row_data.append(cell.value)
            data.append(row_data)
        return data
    else:
        data = []
        for row in ws.iter_rows(values_only=True):
            data.append(list(row))
        return data


def data_to_markdown(data: list, headers: bool = True) -> str:
    """将数据转换为 markdown 表格。"""
    if not data:
        return "*Empty*"
    
    # 计算列宽
    widths = []
    for col_idx in range(len(data[0])):
        max_width = 3
        for row in data:
            if col_idx < len(row):
                val = str(row[col_idx]) if row[col_idx] is not None else ""
                max_width = max(max_width, len(val))
        widths.append(max_width)
    
    lines = []
    for i, row in enumerate(data):
        cells = []
        for j, val in enumerate(row):
            val_str = str(val) if val is not None else ""
            cells.append(val_str.ljust(widths[j]))
        lines.append("| " + " | ".join(cells) + " |")
        
        if i == 0 and headers:
            sep = ["-" * w for w in widths]
            lines.append("| " + " | ".join(sep) + " |")
    
    return "\n".join(lines)


def parse_color(color_str: str) -> str:
    """解析颜色字符串为 ARGB 十六进制。"""
    color_str = color_str.upper().strip()
    
    # 命名颜色
    colors = {
        "RED": "FFFF0000",
        "GREEN": "FF00FF00",
        "BLUE": "FF0000FF",
        "YELLOW": "FFFFFF00",
        "WHITE": "FFFFFFFF",
        "BLACK": "FF000000",
        "GRAY": "FF808080",
        "GREY": "FF808080",
        "ORANGE": "FFFFA500",
        "PURPLE": "FF800080",
        "PINK": "FFFFC0CB",
        "CYAN": "FF00FFFF",
    }
    
    if color_str in colors:
        return colors[color_str]
    
    # 十六进制颜色 (带或不带 #)
    if color_str.startswith("#"):
        color_str = color_str[1:]
    
    if len(color_str) == 6:
        return "FF" + color_str
    elif len(color_str) == 8:
        return color_str
    
    return "FF000000"  # 默认黑色


# ============================================================================
# 命令
# ============================================================================

def cmd_info(args):
    """获取工作簿信息。"""
    require_openpyxl()
    
    try:
        wb = load_workbook(args.file, read_only=False, data_only=True)
    except Exception as e:
        fail(f"Failed to open file: {e}")
    
    sheets_info = []
    for name in wb.sheetnames:
        ws = wb[name]
        sheets_info.append({
            "name": name,
            "dimensions": ws.dimensions or "A1",
            "max_row": ws.max_row,
            "max_column": ws.max_column,
        })
    
    ok({
        "file": args.file,
        "sheets": sheets_info,
        "sheet_count": len(wb.sheetnames),
        "active_sheet": wb.active.title if wb.active else None,
    })
    wb.close()


def cmd_read(args):
    """读取工作表数据。"""
    require_openpyxl()
    
    try:
        wb = load_workbook(args.file, read_only=True, data_only=True)
    except Exception as e:
        fail(f"Failed to open file: {e}")
    
    ws = get_sheet(wb, args.sheet)
    data = read_sheet_data(ws, args.range)
    
    if args.format == "csv":
        import io
        output = io.StringIO()
        writer = csv.writer(output)
        for row in data:
            writer.writerow(row)
        print(output.getvalue())
    elif args.format == "markdown":
        print(data_to_markdown(data))
    else:  # json
        ok({
            "sheet": ws.title,
            "range": args.range or ws.dimensions,
            "rows": len(data),
            "columns": len(data[0]) if data else 0,
            "data": data,
        })
    
    wb.close()


def cmd_cell(args):
    """读取特定单元格。"""
    require_openpyxl()
    
    try:
        wb = load_workbook(args.file, read_only=True, data_only=False)
    except Exception as e:
        fail(f"Failed to open file: {e}")
    
    ws = get_sheet(wb, args.sheet)
    cell = ws[args.cell.upper()]
    
    ok({
        "cell": args.cell.upper(),
        "value": cell.value,
        "formula": cell.value if isinstance(cell.value, str) and cell.value.startswith("=") else None,
        "data_type": cell.data_type,
        "is_merged": cell.coordinate in ws.merged_cells,
    })
    wb.close()


def cmd_create(args):
    """创建新工作簿。"""
    require_openpyxl()
    
    wb = Workbook()
    
    if args.sheets:
        sheet_names = [s.strip() for s in args.sheets.split(",")]
        # 重命名第一个工作表
        wb.active.title = sheet_names[0]
        # 添加其他工作表
        for name in sheet_names[1:]:
            wb.create_sheet(title=name)
    
    try:
        wb.save(args.file)
    except Exception as e:
        fail(f"Failed to save file: {e}")
    
    ok({
        "file": args.file,
        "sheets": wb.sheetnames,
        "created": True,
    })


def cmd_write(args):
    """向单元格写入数据。"""
    require_openpyxl()
    
    # 加载或创建工作簿
    if os.path.exists(args.file):
        wb = load_workbook(args.file)
    else:
        wb = Workbook()
    
    ws = get_sheet(wb, args.sheet)
    
    # 解析数据
    try:
        data = json.loads(args.data)
    except json.JSONDecodeError as e:
        fail(f"Invalid JSON data: {e}")
    
    # 解析起始单元格
    start_row, start_col = cell_to_coords(args.start)
    
    # 写入数据
    if isinstance(data, list):
        for i, row in enumerate(data):
            if isinstance(row, list):
                for j, val in enumerate(row):
                    ws.cell(row=start_row + i, column=start_col + j, value=val)
            else:
                ws.cell(row=start_row + i, column=start_col, value=row)
    elif isinstance(data, dict):
        # 写入为键值对或表头 + 行
        if "headers" in data and "rows" in data:
            for j, header in enumerate(data["headers"]):
                ws.cell(row=start_row, column=start_col + j, value=header)
            for i, row in enumerate(data["rows"]):
                for j, val in enumerate(row):
                    ws.cell(row=start_row + 1 + i, column=start_col + j, value=val)
        else:
            for i, (key, val) in enumerate(data.items()):
                ws.cell(row=start_row + i, column=start_col, value=key)
                ws.cell(row=start_row + i, column=start_col + 1, value=val)
    else:
        ws.cell(row=start_row, column=start_col, value=data)
    
    wb.save(args.file)
    ok({
        "file": args.file,
        "sheet": ws.title,
        "start": args.start,
        "written": True,
    })


def cmd_from_csv(args):
    """从 CSV 创建 Excel。"""
    require_openpyxl()
    
    wb = Workbook()
    ws = wb.active
    if args.sheet:
        ws.title = args.sheet
    
    try:
        with open(args.csv_file, 'r', newline='', encoding='utf-8-sig') as f:
            reader = csv.reader(f)
            for row in reader:
                ws.append(row)
    except Exception as e:
        fail(f"Failed to read CSV: {e}")
    
    wb.save(args.excel_file)
    ok({
        "source": args.csv_file,
        "output": args.excel_file,
        "sheet": ws.title,
        "rows": ws.max_row,
    })


def cmd_from_json(args):
    """从 JSON 创建 Excel。"""
    require_openpyxl()
    
    try:
        with open(args.json_file, 'r') as f:
            data = json.load(f)
    except Exception as e:
        fail(f"Failed to read JSON: {e}")
    
    wb = Workbook()
    ws = wb.active
    if args.sheet:
        ws.title = args.sheet
    
    if isinstance(data, list):
        if data and isinstance(data[0], dict):
            # 字典列表 - 使用键作为表头
            headers = list(data[0].keys())
            ws.append(headers)
            for item in data:
                ws.append([item.get(h) for h in headers])
        else:
            # 列表的列表或值
            for row in data:
                if isinstance(row, list):
                    ws.append(row)
                else:
                    ws.append([row])
    elif isinstance(data, dict):
        if "headers" in data and "rows" in data:
            ws.append(data["headers"])
            for row in data["rows"]:
                ws.append(row)
        else:
            for key, val in data.items():
                ws.append([key, val])
    
    wb.save(args.excel_file)
    ok({
        "source": args.json_file,
        "output": args.excel_file,
        "sheet": ws.title,
        "rows": ws.max_row,
    })


def cmd_edit(args):
    """编辑单元格值或公式。"""
    require_openpyxl()
    
    wb = load_workbook(args.file)
    ws = get_sheet(wb, args.sheet)
    
    cell = ws[args.cell.upper()]
    old_value = cell.value
    
    # 设置新值
    if args.formula:
        if not args.value.startswith("="):
            args.value = "=" + args.value
        cell.value = args.value
    else:
        # 尝试转换为适当类型
        try:
            if args.value.lower() in ('true', 'false'):
                cell.value = args.value.lower() == 'true'
            elif '.' in args.value:
                cell.value = float(args.value)
            else:
                cell.value = int(args.value)
        except ValueError:
            cell.value = args.value
    
    wb.save(args.file)
    ok({
        "cell": args.cell.upper(),
        "old_value": old_value,
        "new_value": cell.value,
        "is_formula": args.formula,
    })


def cmd_add_sheet(args):
    """添加新工作表。"""
    require_openpyxl()
    
    wb = load_workbook(args.file)
    
    if args.name in wb.sheetnames:
        fail(f"Sheet '{args.name}' already exists")
    
    if args.position is not None:
        wb.create_sheet(title=args.name, index=args.position)
    else:
        wb.create_sheet(title=args.name)
    
    wb.save(args.file)
    ok({
        "added": args.name,
        "sheets": wb.sheetnames,
    })


def cmd_rename_sheet(args):
    """重命名工作表。"""
    require_openpyxl()
    
    wb = load_workbook(args.file)
    
    if args.old_name not in wb.sheetnames:
        fail(f"Sheet '{args.old_name}' not found")
    
    if args.new_name in wb.sheetnames:
        fail(f"Sheet '{args.new_name}' already exists")
    
    wb[args.old_name].title = args.new_name
    wb.save(args.file)
    
    ok({
        "old_name": args.old_name,
        "new_name": args.new_name,
        "sheets": wb.sheetnames,
    })


def cmd_delete_sheet(args):
    """删除工作表。"""
    require_openpyxl()
    
    wb = load_workbook(args.file)
    
    if args.name not in wb.sheetnames:
        fail(f"Sheet '{args.name}' not found")
    
    if len(wb.sheetnames) == 1:
        fail("Cannot delete the only sheet in workbook")
    
    del wb[args.name]
    wb.save(args.file)
    
    ok({
        "deleted": args.name,
        "sheets": wb.sheetnames,
    })


def cmd_copy_sheet(args):
    """复制工作表。"""
    require_openpyxl()
    
    wb = load_workbook(args.file)
    
    if args.source not in wb.sheetnames:
        fail(f"Sheet '{args.source}' not found")
    
    if args.new_name in wb.sheetnames:
        fail(f"Sheet '{args.new_name}' already exists")
    
    source_ws = wb[args.source]
    new_ws = wb.copy_worksheet(source_ws)
    new_ws.title = args.new_name
    
    wb.save(args.file)
    ok({
        "source": args.source,
        "copy": args.new_name,
        "sheets": wb.sheetnames,
    })


def cmd_insert_rows(args):
    """插入行。"""
    require_openpyxl()
    
    wb = load_workbook(args.file)
    ws = get_sheet(wb, args.sheet)
    
    ws.insert_rows(args.row, args.count)
    wb.save(args.file)
    
    ok({
        "inserted": "rows",
        "at": args.row,
        "count": args.count,
        "sheet": ws.title,
    })


def cmd_insert_cols(args):
    """插入列。"""
    require_openpyxl()
    
    wb = load_workbook(args.file)
    ws = get_sheet(wb, args.sheet)
    
    # 如有需要，将列字母转换为数字
    col = args.col
    if isinstance(col, str) and col.isalpha():
        col = column_index_from_string(col.upper())
    
    ws.insert_cols(int(col), args.count)
    wb.save(args.file)
    
    ok({
        "inserted": "columns",
        "at": args.col,
        "count": args.count,
        "sheet": ws.title,
    })


def cmd_delete_rows(args):
    """删除行。"""
    require_openpyxl()
    
    wb = load_workbook(args.file)
    ws = get_sheet(wb, args.sheet)
    
    ws.delete_rows(args.row, args.count)
    wb.save(args.file)
    
    ok({
        "deleted": "rows",
        "at": args.row,
        "count": args.count,
        "sheet": ws.title,
    })


def cmd_delete_cols(args):
    """删除列。"""
    require_openpyxl()
    
    wb = load_workbook(args.file)
    ws = get_sheet(wb, args.sheet)
    
    col = args.col
    if isinstance(col, str) and col.isalpha():
        col = column_index_from_string(col.upper())
    
    ws.delete_cols(int(col), args.count)
    wb.save(args.file)
    
    ok({
        "deleted": "columns",
        "at": args.col,
        "count": args.count,
        "sheet": ws.title,
    })


def cmd_merge(args):
    """合并单元格。"""
    require_openpyxl()
    
    wb = load_workbook(args.file)
    ws = get_sheet(wb, args.sheet)
    
    ws.merge_cells(args.range)
    wb.save(args.file)
    
    ok({
        "merged": args.range,
        "sheet": ws.title,
    })


def cmd_unmerge(args):
    """取消合并单元格。"""
    require_openpyxl()
    
    wb = load_workbook(args.file)
    ws = get_sheet(wb, args.sheet)
    
    ws.unmerge_cells(args.range)
    wb.save(args.file)
    
    ok({
        "unmerged": args.range,
        "sheet": ws.title,
    })


def cmd_format(args):
    """格式化单元格。"""
    require_openpyxl()
    
    wb = load_workbook(args.file)
    ws = get_sheet(wb, args.sheet)
    
    (start_row, start_col), (end_row, end_col) = parse_range(args.range)
    
    applied = []
    
    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            cell = ws.cell(row=row, column=col)
            
            # 字体
            if args.bold is not None or args.italic is not None or args.font_size or args.font_color or args.font_name:
                current_font = cell.font
                cell.font = Font(
                    name=args.font_name or current_font.name,
                    size=args.font_size or current_font.size,
                    bold=args.bold if args.bold is not None else current_font.bold,
                    italic=args.italic if args.italic is not None else current_font.italic,
                    color=parse_color(args.font_color) if args.font_color else current_font.color,
                )
                applied.append("font")
            
            # 填充/背景
            if args.bg_color:
                cell.fill = PatternFill(start_color=parse_color(args.bg_color),
                                         end_color=parse_color(args.bg_color),
                                         fill_type="solid")
                applied.append("background")
            
            # 对齐
            if args.align or args.valign or args.wrap:
                current_align = cell.alignment
                cell.alignment = Alignment(
                    horizontal=args.align or current_align.horizontal,
                    vertical=args.valign or current_align.vertical,
                    wrap_text=args.wrap if args.wrap is not None else current_align.wrap_text,
                )
                applied.append("alignment")
            
            # 边框
            if args.border:
                side = Side(style=args.border, color="FF000000")
                cell.border = Border(left=side, right=side, top=side, bottom=side)
                applied.append("border")
    
    wb.save(args.file)
    ok({
        "range": args.range,
        "sheet": ws.title,
        "applied": list(set(applied)),
    })


def cmd_resize(args):
    """调整行和列大小。"""
    require_openpyxl()
    
    wb = load_workbook(args.file)
    ws = get_sheet(wb, args.sheet)
    
    resized = []
    
    if args.row:
        for spec in args.row:
            row_num, height = spec.split(':')
            ws.row_dimensions[int(row_num)].height = float(height)
            resized.append(f"row {row_num} = {height}")
    
    if args.col:
        for spec in args.col:
            col_letter, width = spec.split(':')
            ws.column_dimensions[col_letter.upper()].width = float(width)
            resized.append(f"col {col_letter} = {width}")
    
    wb.save(args.file)
    ok({
        "sheet": ws.title,
        "resized": resized,
    })


def cmd_freeze(args):
    """冻结窗格。"""
    require_openpyxl()
    
    wb = load_workbook(args.file)
    ws = get_sheet(wb, args.sheet)
    
    ws.freeze_panes = args.cell.upper()
    wb.save(args.file)
    
    ok({
        "sheet": ws.title,
        "frozen_at": args.cell.upper(),
    })


def cmd_find(args):
    """在工作表中查找文本。"""
    require_openpyxl()
    
    wb = load_workbook(args.file, read_only=True, data_only=True)
    ws = get_sheet(wb, args.sheet)
    
    results = []
    search_lower = args.text.lower()
    
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and search_lower in str(cell.value).lower():
                results.append({
                    "cell": cell.coordinate,
                    "value": cell.value,
                })
    
    ok({
        "search": args.text,
        "sheet": ws.title,
        "found": len(results),
        "results": results,
    })
    wb.close()


def cmd_replace(args):
    """查找并替换文本。"""
    require_openpyxl()
    
    wb = load_workbook(args.file)
    ws = get_sheet(wb, args.sheet)
    
    count = 0
    
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str) and args.old in cell.value:
                cell.value = cell.value.replace(args.old, args.new)
                count += 1
    
    wb.save(args.file)
    ok({
        "old": args.old,
        "new": args.new,
        "sheet": ws.title,
        "replaced": count,
    })


def cmd_to_csv(args):
    """将工作表导出为 CSV。"""
    require_openpyxl()
    
    wb = load_workbook(args.file, read_only=True, data_only=True)
    ws = get_sheet(wb, args.sheet)
    
    with open(args.output, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        for row in ws.iter_rows(values_only=True):
            writer.writerow(row)
    
    ok({
        "source": args.file,
        "sheet": ws.title,
        "output": args.output,
        "rows": ws.max_row,
    })
    wb.close()


def cmd_to_json(args):
    """将工作表导出为 JSON。"""
    require_openpyxl()
    
    wb = load_workbook(args.file, read_only=True, data_only=True)
    ws = get_sheet(wb, args.sheet)
    
    data = []
    headers = None
    
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            headers = list(row)
        else:
            row_dict = {}
            for j, val in enumerate(row):
                if j < len(headers) and headers[j]:
                    row_dict[headers[j]] = val
            data.append(row_dict)
    
    with open(args.output, 'w', encoding='utf-8') as f:
        json.dump(data, f, indent=2, default=str)
    
    ok({
        "source": args.file,
        "sheet": ws.title,
        "output": args.output,
        "rows": len(data),
    })
    wb.close()


def cmd_to_markdown(args):
    """将工作表导出为 markdown 表格。"""
    require_openpyxl()
    
    wb = load_workbook(args.file, read_only=True, data_only=True)
    ws = get_sheet(wb, args.sheet)
    
    data = read_sheet_data(ws)
    print(data_to_markdown(data))
    wb.close()


def main():
    parser = argparse.ArgumentParser(
        description="Excel 命令行工具 - 读取、写入、编辑和格式化 Excel 文件",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    subparsers = parser.add_subparsers(dest="command", help="命令")
    
    # info
    p = subparsers.add_parser("info", help="获取工作簿信息")
    p.add_argument("file", help="Excel 文件路径")
    
    # read
    p = subparsers.add_parser("read", help="读取工作表数据")
    p.add_argument("file", help="Excel 文件路径")
    p.add_argument("--sheet", "-s", help="工作表名称")
    p.add_argument("--range", "-r", help="单元格范围（例如：A1:D10）")
    p.add_argument("--format", "-f", choices=["json", "csv", "markdown"], default="json")
    
    # cell
    p = subparsers.add_parser("cell", help="读取特定单元格")
    p.add_argument("file", help="Excel 文件路径")
    p.add_argument("cell", help="单元格引用（例如：A1）")
    p.add_argument("--sheet", "-s", help="工作表名称")
    
    # create
    p = subparsers.add_parser("create", help="创建新工作簿")
    p.add_argument("file", help="输出文件路径")
    p.add_argument("--sheets", help="逗号分隔的工作表名称")
    
    # write
    p = subparsers.add_parser("write", help="向单元格写入数据")
    p.add_argument("file", help="Excel 文件路径")
    p.add_argument("--data", "-d", required=True, help="要写入的 JSON 数据")
    p.add_argument("--sheet", "-s", help="工作表名称")
    p.add_argument("--start", default="A1", help="起始单元格（默认：A1）")
    
    # from-csv
    p = subparsers.add_parser("from-csv", help="从 CSV 创建 Excel")
    p.add_argument("csv_file", help="输入 CSV 文件")
    p.add_argument("excel_file", help="输出 Excel 文件")
    p.add_argument("--sheet", "-s", help="工作表名称")
    
    # from-json
    p = subparsers.add_parser("from-json", help="从 JSON 创建 Excel")
    p.add_argument("json_file", help="输入 JSON 文件")
    p.add_argument("excel_file", help="输出 Excel 文件")
    p.add_argument("--sheet", "-s", help="工作表名称")
    
    # edit
    p = subparsers.add_parser("edit", help="编辑单元格")
    p.add_argument("file", help="Excel 文件路径")
    p.add_argument("cell", help="单元格引用")
    p.add_argument("value", help="新值")
    p.add_argument("--sheet", "-s", help="工作表名称")
    p.add_argument("--formula", "-F", action="store_true", help="值是公式")
    
    # add-sheet
    p = subparsers.add_parser("add-sheet", help="添加新工作表")
    p.add_argument("file", help="Excel 文件路径")
    p.add_argument("name", help="新工作表名称")
    p.add_argument("--position", "-p", type=int, help="位置索引")
    
    # rename-sheet
    p = subparsers.add_parser("rename-sheet", help="重命名工作表")
    p.add_argument("file", help="Excel 文件路径")
    p.add_argument("old_name", help="当前工作表名称")
    p.add_argument("new_name", help="新工作表名称")
    
    # delete-sheet
    p = subparsers.add_parser("delete-sheet", help="删除工作表")
    p.add_argument("file", help="Excel 文件路径")
    p.add_argument("name", help="要删除的工作表名称")
    
    # copy-sheet
    p = subparsers.add_parser("copy-sheet", help="复制工作表")
    p.add_argument("file", help="Excel 文件路径")
    p.add_argument("source", help="源工作表名称")
    p.add_argument("new_name", help="新工作表名称")
    
    # insert-rows
    p = subparsers.add_parser("insert-rows", help="插入行")
    p.add_argument("file", help="Excel 文件路径")
    p.add_argument("row", type=int, help="要插入的行号")
    p.add_argument("--count", "-n", type=int, default=1, help="行数")
    p.add_argument("--sheet", "-s", help="工作表名称")
    
    # insert-cols
    p = subparsers.add_parser("insert-cols", help="插入列")
    p.add_argument("file", help="Excel 文件路径")
    p.add_argument("col", help="要插入的列（字母或数字）")
    p.add_argument("--count", "-n", type=int, default=1, help="列数")
    p.add_argument("--sheet", "-s", help="工作表名称")
    
    # delete-rows
    p = subparsers.add_parser("delete-rows", help="删除行")
    p.add_argument("file", help="Excel 文件路径")
    p.add_argument("row", type=int, help="起始行号")
    p.add_argument("--count", "-n", type=int, default=1, help="行数")
    p.add_argument("--sheet", "-s", help="工作表名称")
    
    # delete-cols
    p = subparsers.add_parser("delete-cols", help="删除列")
    p.add_argument("file", help="Excel 文件路径")
    p.add_argument("col", help="列（字母或数字）")
    p.add_argument("--count", "-n", type=int, default=1, help="列数")
    p.add_argument("--sheet", "-s", help="工作表名称")
    
    # merge
    p = subparsers.add_parser("merge", help="合并单元格")
    p.add_argument("file", help="Excel 文件路径")
    p.add_argument("range", help="要合并的单元格范围（例如：A1:C1）")
    p.add_argument("--sheet", "-s", help="工作表名称")
    
    # unmerge
    p = subparsers.add_parser("unmerge", help="取消合并单元格")
    p.add_argument("file", help="Excel 文件路径")
    p.add_argument("range", help="要取消合并的单元格范围")
    p.add_argument("--sheet", "-s", help="工作表名称")
    
    # format
    p = subparsers.add_parser("format", help="格式化单元格")
    p.add_argument("file", help="Excel 文件路径")
    p.add_argument("range", help="要格式化的单元格范围")
    p.add_argument("--sheet", "-s", help="工作表名称")
    p.add_argument("--bold", "-b", action="store_true", default=None)
    p.add_argument("--no-bold", dest="bold", action="store_false")
    p.add_argument("--italic", "-i", action="store_true", default=None)
    p.add_argument("--no-italic", dest="italic", action="store_false")
    p.add_argument("--font-size", type=int, help="字体大小")
    p.add_argument("--font-color", help="字体颜色（名称或十六进制）")
    p.add_argument("--font-name", help="字体名称")
    p.add_argument("--bg-color", help="背景颜色")
    p.add_argument("--align", choices=["left", "center", "right"])
    p.add_argument("--valign", choices=["top", "center", "bottom"])
    p.add_argument("--wrap", action="store_true", default=None)
    p.add_argument("--no-wrap", dest="wrap", action="store_false")
    p.add_argument("--border", choices=["thin", "medium", "thick", "double"])
    
    # resize
    p = subparsers.add_parser("resize", help="调整行和列大小")
    p.add_argument("file", help="Excel 文件路径")
    p.add_argument("--row", action="append", help="Row:height（例如：1:30）")
    p.add_argument("--col", action="append", help="Col:width（例如：A:20）")
    p.add_argument("--sheet", "-s", help="工作表名称")
    
    # freeze
    p = subparsers.add_parser("freeze", help="冻结窗格")
    p.add_argument("file", help="Excel 文件路径")
    p.add_argument("cell", help="要冻结的单元格（例如：A2 冻结第 1 行）")
    p.add_argument("--sheet", "-s", help="工作表名称")
    
    # find
    p = subparsers.add_parser("find", help="在工作表中查找文本")
    p.add_argument("file", help="Excel 文件路径")
    p.add_argument("text", help="要搜索的文本")
    p.add_argument("--sheet", "-s", help="工作表名称")
    
    # replace
    p = subparsers.add_parser("replace", help="查找并替换文本")
    p.add_argument("file", help="Excel 文件路径")
    p.add_argument("old", help="要查找的文本")
    p.add_argument("new", help="替换文本")
    p.add_argument("--sheet", "-s", help="工作表名称")
    
    # to-csv
    p = subparsers.add_parser("to-csv", help="将工作表导出为 CSV")
    p.add_argument("file", help="Excel 文件路径")
    p.add_argument("output", help="输出 CSV 文件")
    p.add_argument("--sheet", "-s", help="工作表名称")
    
    # to-json
    p = subparsers.add_parser("to-json", help="将工作表导出为 JSON")
    p.add_argument("file", help="Excel 文件路径")
    p.add_argument("output", help="输出 JSON 文件")
    p.add_argument("--sheet", "-s", help="工作表名称")
    
    # to-markdown
    p = subparsers.add_parser("to-markdown", help="将工作表导出为 markdown")
    p.add_argument("file", help="Excel 文件路径")
    p.add_argument("--sheet", "-s", help="工作表名称")
    
    args = parser.parse_args()
    
    if not args.command:
        parser.print_help()
        sys.exit(1)
    
    commands = {
        "info": cmd_info,
        "read": cmd_read,
        "cell": cmd_cell,
        "create": cmd_create,
        "write": cmd_write,
        "from-csv": cmd_from_csv,
        "from-json": cmd_from_json,
        "edit": cmd_edit,
        "add-sheet": cmd_add_sheet,
        "rename-sheet": cmd_rename_sheet,
        "delete-sheet": cmd_delete_sheet,
        "copy-sheet": cmd_copy_sheet,
        "insert-rows": cmd_insert_rows,
        "insert-cols": cmd_insert_cols,
        "delete-rows": cmd_delete_rows,
        "delete-cols": cmd_delete_cols,
        "merge": cmd_merge,
        "unmerge": cmd_unmerge,
        "format": cmd_format,
        "resize": cmd_resize,
        "freeze": cmd_freeze,
        "find": cmd_find,
        "replace": cmd_replace,
        "to-csv": cmd_to_csv,
        "to-json": cmd_to_json,
        "to-markdown": cmd_to_markdown,
    }
    
    try:
        commands[args.command](args)
    except Exception as e:
        fail(str(e))


if __name__ == "__main__":
    main()